#!/usr/bin/env python3
"""
plot_figure2A.py

Reproduce Figure 2A of Kelliher et al. (2016, PLOS Genetics,
doi:10.1371/journal.pgen.1006453): a heatmap of the 1246 periodic
S. cerevisiae genes during the cell cycle, ordered by peak expression
time, with transcript levels shown as z-scores.

Usage
-----
    python plot_figure2A.py <path_to_S1_Table.xlsx> [-o output.png]

Input
-----
The input file is the S1 Table supplementary spreadsheet
(pgen_1006453_s002.xlsx), containing periodicity rankings and the
RNA-Seq time series (FPKM) for S. cerevisiae.

Relevant columns (header row 3 of the sheet):
    - gene_ID                   : gene identifier
    - normalized_per_rank       : cumulative periodicity rank (non-noisy genes)
    - LS_cutoff                 : "Yes"/"No", Lomb-Scargle score cutoff flag
    - Figure2A_order_peaktime   : y-axis order used in Figure 2A
    - time-series columns (0, 5, 10, ..., 245 minutes) : FPKM values
"""

from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl


HEADER_ROW = 3
DATA_START_ROW = 4

COL_GENE_ID = "gene_ID"
COL_NORMALIZED_RANK = "normalized_per_rank"
COL_LS_CUTOFF = "LS_cutoff"
COL_FIG2A_ORDER = "Figure2A_order_peaktime"

TOP_N_CUMULATIVE_RANK = 1600  # top 1600 cumulative periodicity ranking


def load_periodic_gene_matrix(xlsx_path: Path) -> tuple[np.ndarray, list[str], np.ndarray]:
    """Load the periodic gene expression matrix used in Figure 2A.

    Parameters
    ----------
    xlsx_path:
        Path to the S1 Table supplementary spreadsheet.

    Returns
    -------
    matrix:
        Array of shape (n_genes, n_timepoints) with raw FPKM values,
        ordered by ``Figure2A_order_peaktime``.
    gene_ids:
        Gene identifiers, in the same order as ``matrix`` rows.
    time_points:
        Array of time points (minutes) corresponding to the columns
        of ``matrix``.
    """
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]

    col_idx = {name: i for i, name in enumerate(header) if name is not None}

    # Time-series columns are the numeric headers (0, 5, 10, ..., 245)
    time_cols = [
        (i, value) for i, value in enumerate(header) if isinstance(value, (int, float))
    ]
    time_cols.sort(key=lambda pair: pair[1])
    time_indices = [i for i, _ in time_cols]
    time_points = np.array([v for _, v in time_cols], dtype=float)

    records = []
    for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        ls_cutoff = row[col_idx[COL_LS_CUTOFF]]
        rank = row[col_idx[COL_NORMALIZED_RANK]]
        order = row[col_idx[COL_FIG2A_ORDER]]

        if ls_cutoff != "Yes":
            continue
        if not isinstance(rank, (int, float)) or rank > TOP_N_CUMULATIVE_RANK:
            continue
        if order is None:
            continue

        gene_id = row[col_idx[COL_GENE_ID]]
        values = [row[i] for i in time_indices]
        records.append((order, gene_id, values))

    workbook.close()

    if not records:
        raise ValueError("No periodic genes found - check input file and column names.")

    records.sort(key=lambda r: r[0])

    gene_ids = [gene_id for _, gene_id, _ in records]
    matrix = np.array([values for _, _, values in records], dtype=float)

    return matrix, gene_ids, time_points


def to_zscore(matrix: np.ndarray) -> np.ndarray:
    """Convert each row of ``matrix`` (one gene) to a z-score profile."""
    mean = matrix.mean(axis=1, keepdims=True)
    std = matrix.std(axis=1, ddof=0, keepdims=True)
    return (matrix - mean) / std


def plot_heatmap(zscores: np.ndarray, time_points: np.ndarray, output_path: Path) -> None:
    """Plot the periodic gene expression heatmap (Figure 2A style)."""
    fig, ax = plt.subplots(figsize=(6, 8))

    extent = (time_points[0], time_points[-1], zscores.shape[0], 0)
    im = ax.imshow(
        zscores,
        aspect="auto",
        cmap="cividis",
        vmin=-1.5,
        vmax=1.5,
        extent=extent,
        interpolation="nearest",
    )

    ax.set_xlabel("time (minutes)")
    ax.set_ylabel(f"Top Periodic Genes ({zscores.shape[0]})")
    ax.set_title("Saccharomyces cerevisiae")

    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.set_label("z-score")

    fig.tight_layout()
    fig.savefig(output_path, dpi=300)
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", type=Path, help="Path to S1 Table (.xlsx)")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("figure2A_reproduction.png"),
        help="Output image path (default: figure2A_reproduction.png)",
    )
    args = parser.parse_args()

    matrix, gene_ids, time_points = load_periodic_gene_matrix(args.xlsx_path)
    zscores = to_zscore(matrix)

    print(f"Loaded {len(gene_ids)} periodic genes x {len(time_points)} time points.")
    plot_heatmap(zscores, time_points, args.output)
    print(f"Figure saved to: {args.output}")


if __name__ == "__main__":
    main()