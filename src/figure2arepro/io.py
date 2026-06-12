from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl

HEADER_ROW = 3
DATA_START_ROW = 4

COL_GENE_ID = "gene_ID"
COL_NORMALIZED_RANK = "normalized_per_rank"
COL_LS_CUTOFF = "LS_cutoff"
COL_FIG2A_ORDER = "Figure2A_order_peaktime"
TOP_N_CUMULATIVE_RANK = 1600


def load_periodic_gene_matrix(xlsx_path: Path) -> tuple[np.ndarray, list[str], np.ndarray]:
    """Load the periodic gene expression matrix used in Figure 2A.

    Parameters
    ----------
    xlsx_path:
        Path to the S1 Table supplementary spreadsheet.

    Returns
    -------
    matrix:
        Array of shape (n_genes, n_timepoints) with raw FPKM values,
        ordered by ``Figure2A_order_peaktime``.
    gene_ids:
        Gene identifiers, in the same order as ``matrix`` rows.
    time_points:
        Array of time points (minutes) corresponding to the columns
        of ``matrix``.
    """
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    header = [cell.value for cell in next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
    col_idx = {name: i for i, name in enumerate(header) if name is not None}

    time_cols = [(i, value) for i, value in enumerate(header) if isinstance(value, (int, float))]
    time_cols.sort(key=lambda pair: pair[1])
    time_indices = [i for i, _ in time_cols]
    time_points = np.array([v for _, v in time_cols], dtype=float)

    records: list[tuple[float, str, list[float]]] = []
    for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        ls_cutoff = row[col_idx[COL_LS_CUTOFF]]
        rank = row[col_idx[COL_NORMALIZED_RANK]]
        order = row[col_idx[COL_FIG2A_ORDER]]

        if ls_cutoff != "Yes":
            continue
        if not isinstance(rank, (int, float)) or rank > TOP_N_CUMULATIVE_RANK:
            continue
        if order is None:
            continue

        gene_id = row[col_idx[COL_GENE_ID]]
        if gene_id is None:
            continue
        if not isinstance(gene_id, str):
            gene_id = str(gene_id)

        values = [row[i] for i in time_indices]
        records.append((order, gene_id, values))

    workbook.close()

    if not records:
        raise ValueError("No periodic genes found - check input file and column names.")

    records.sort(key=lambda r: r[0])

    gene_ids = [gene_id for _, gene_id, _ in records]
    matrix = np.array([values for _, _, values in records], dtype=float)

    return matrix, gene_ids, time_points
